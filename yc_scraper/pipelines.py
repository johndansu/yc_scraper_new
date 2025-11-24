# Define your item pipelines here
#
# Don't forget to add your pipeline to the ITEM_PIPELINES setting
# See: https://docs.scrapy.org/en/latest/topics/item-pipeline.html

import pandas as pd
from itemadapter import ItemAdapter
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter
import re


class ExcelExportPipeline:
    """Pipeline to export items to Excel format"""

    def __init__(self):
        self.items = []
        self.original_urls = []  # Store original URLs for hyperlinks

    def process_item(self, item, spider):
        adapter = ItemAdapter(item)
        
        # Clean and format each field
        company_name = adapter.get('company_name', '').strip()
        
        # Format company website to short format (www.example.com)
        company_website_raw = adapter.get('company_website', '').strip()
        company_website = self._format_website(company_website_raw)
        
        # Clean founder names - remove any URLs, keep only text
        founders_name_raw = adapter.get('founders_name', '').strip()
        founders_name = self._clean_founder_names(founders_name_raw)
        
        # Format LinkedIn links to short format
        founders_linkedin_raw = adapter.get('founders_linkedin', '').strip()
        founders_linkedin = self._format_linkedin(founders_linkedin_raw)
        
        # Format Twitter links to short format
        founders_twitter_raw = adapter.get('founders_twitter', '').strip()
        founders_twitter = self._format_twitter(founders_twitter_raw)
        
        item_dict = {
            'company_name': company_name,
            'company_website': company_website,
            'founders_name': founders_name,
            'founders_linkedin': founders_linkedin,
            'founders_twitter': founders_twitter
        }
        self.items.append(item_dict)
        
        # Store original URLs for hyperlink creation
        self.original_urls.append({
            'company_website': company_website_raw,
            'founders_linkedin': founders_linkedin_raw,
            'founders_twitter': founders_twitter_raw
        })
        
        return item

    def close_spider(self, spider):
        if not self.items:
            spider.logger.warning('No items to export')
            return
        
        # Create DataFrame with proper column names
        df = pd.DataFrame(self.items)
        
        # Rename columns to user-friendly names
        column_mapping = {
            'company_name': 'Company Name',
            'company_website': 'Company Website',
            'founders_name': "Founder's Name",
            'founders_linkedin': 'Founders LinkedIn',
            'founders_twitter': 'Founders Twitter'
        }
        
        df = df.rename(columns=column_mapping)
        
        # Ensure columns are in the correct order
        column_order = [
            'Company Name',
            'Company Website',
            "Founder's Name",
            'Founders LinkedIn',
            'Founders Twitter'
        ]
        
        # Reorder columns (only include columns that exist)
        df = df[[col for col in column_order if col in df.columns]]
        
        # Export to Excel
        output_file = 'yc_companies.xlsx'
        
        try:
            # Write to Excel
            df.to_excel(output_file, index=False, engine='openpyxl')
            
            # Format the Excel file for better readability
            self._format_excel_file(output_file)
            
            spider.logger.info(f'✅ Successfully exported {len(self.items)} companies to {output_file}')
        except Exception as e:
            spider.logger.error(f'Error exporting to Excel: {e}')
            raise
    
    def _format_website(self, url):
        """Format website URL to short format: www.example.com, exclude unwanted domains"""
        if not url:
            return ''
        
        # Exclude unwanted domains
        excluded_domains = [
            'startupschool.org', 'startupschool.com', 'ycombinator.com',
            'bookface-static.ycombinator.com', 'bookface-images.s3'
        ]
        
        # Check if URL contains excluded domains
        url_lower = url.lower()
        if any(excluded in url_lower for excluded in excluded_domains):
            return ''  # Return empty if it's an excluded domain
        
        try:
            # Extract domain from URL
            match = re.search(r'https?://(?:www\.)?([^/?#\s]+)', url)
            if match:
                domain = match.group(1)
                # Check again for excluded domains in the extracted domain
                if any(excluded in domain.lower() for excluded in excluded_domains):
                    return ''
                # Add www. prefix if not present
                if not domain.startswith('www.'):
                    domain = 'www.' + domain
                return domain
        except:
            pass
        
        # If already looks like a domain, check and format
        if url and not url.startswith('http') and '.' in url:
            if any(excluded in url.lower() for excluded in excluded_domains):
                return ''
            if not url.startswith('www.'):
                return 'www.' + url
            return url
        
        return url
    
    def _clean_founder_names(self, names_text):
        """Clean founder names - remove URLs, keep only text names"""
        if not names_text:
            return ''
        
        # Remove any URLs that might be mixed in
        # Remove patterns like http://, https://, linkedin.com, twitter.com, etc.
        cleaned = re.sub(r'https?://[^\s,]+', '', names_text)
        cleaned = re.sub(r'linkedin\.com/[^\s,]+', '', cleaned)
        cleaned = re.sub(r'twitter\.com/[^\s,]+', '', cleaned)
        cleaned = re.sub(r'x\.com/[^\s,]+', '', cleaned)
        
        # Remove multiple spaces and clean up
        cleaned = re.sub(r'\s+', ' ', cleaned)
        cleaned = re.sub(r',\s*,', ',', cleaned)  # Remove double commas
        cleaned = cleaned.strip(' ,')
        
        return cleaned
    
    def _format_linkedin(self, linkedin_text):
        """Format LinkedIn URLs to short format: linkedin.com/in/username"""
        if not linkedin_text:
            return ''
        
        # Handle multiple LinkedIn URLs separated by commas
        if ',' in linkedin_text:
            urls = [u.strip() for u in linkedin_text.split(',')]
            formatted = [self._format_single_linkedin(url) for url in urls if url]
            return ', '.join([f for f in formatted if f])
        
        return self._format_single_linkedin(linkedin_text)
    
    def _format_single_linkedin(self, url):
        """Format a single LinkedIn URL"""
        if not url:
            return ''
        
        try:
            # Extract LinkedIn username: linkedin.com/in/username
            match = re.search(r'linkedin\.com/in/([^/?\s]+)', url)
            if match:
                return f"linkedin.com/in/{match.group(1)}"
        except:
            pass
        
        return url
    
    def _format_twitter(self, twitter_text):
        """Format Twitter URLs to short format: @username, filter out @ycombinator"""
        if not twitter_text:
            return ''
        
        # Handle multiple Twitter URLs separated by commas
        if ',' in twitter_text:
            urls = [u.strip() for u in twitter_text.split(',')]
            formatted = [self._format_single_twitter(url) for url in urls if url]
            # Filter out @ycombinator
            filtered = [f for f in formatted if f and 'ycombinator' not in f.lower()]
            return ', '.join(filtered) if filtered else ''
        
        result = self._format_single_twitter(twitter_text)
        # Filter out @ycombinator
        if result and 'ycombinator' in result.lower():
            return ''
        return result
    
    def _format_single_twitter(self, url):
        """Format a single Twitter URL"""
        if not url:
            return ''
        
        try:
            # Extract Twitter/X username: twitter.com/username or x.com/username
            match = re.search(r'(?:twitter|x)\.com/([^/?\s]+)', url)
            if match:
                username = match.group(1)
                # Remove @ if present
                username = username.lstrip('@')
                return f"@{username}"
        except:
            pass
        
        return url
    
    def _format_excel_file(self, filename):
        """Format the Excel file with headers, hyperlinks, and column widths"""
        try:
            wb = load_workbook(filename)
            ws = wb.active
            
            # Format header row
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF", size=11)
            link_font = Font(color="0000FF", underline="single")
            
            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center', vertical='center')
            
            # Store original URLs for hyperlinks (from items list)
            # Process each row and add hyperlinks where appropriate
            for row_idx, row in enumerate(ws.iter_rows(min_row=2, max_row=ws.max_row), start=2):
                item_idx = row_idx - 2  # Adjust for 0-based index and header row
                if item_idx < len(self.original_urls):
                    original_urls = self.original_urls[item_idx]
                    
                    for col_idx, cell in enumerate(row, start=1):
                        if cell.value:
                            cell_value = str(cell.value).strip()
                            
                            # Get column header to determine link type
                            header_cell = ws.cell(row=1, column=col_idx)
                            column_name = str(header_cell.value).strip() if header_cell.value else ''
                            
                            # Add hyperlinks to clickable fields
                            if column_name == 'Company Website' and cell_value and 'www.' in cell_value:
                                # Create hyperlink from www.example.com to https://www.example.com
                                original_url = original_urls.get('company_website', '')
                                if original_url and 'http' in original_url.lower():
                                    url = original_url.split(',')[0].strip()
                                    cell.hyperlink = url
                                    cell.font = link_font
                                elif cell_value and 'www.' in cell_value:
                                    url = 'https://' + cell_value if not cell_value.startswith('http') else cell_value
                                    cell.hyperlink = url
                                    cell.font = link_font
                            
                            elif column_name == 'Founders LinkedIn' and cell_value:
                                # Extract full URL from original data
                                original_linkedin = original_urls.get('founders_linkedin', '')
                                if original_linkedin and 'http' in original_linkedin.lower():
                                    # Use first URL if multiple
                                    url = original_linkedin.split(',')[0].strip()
                                    cell.hyperlink = url
                                    cell.font = link_font
                            
                            elif column_name == 'Founders Twitter' and cell_value:
                                # Extract full URL from original data
                                original_twitter = original_urls.get('founders_twitter', '')
                                if original_twitter and 'http' in original_twitter.lower():
                                    # Use first URL if multiple
                                    url = original_twitter.split(',')[0].strip()
                                    cell.hyperlink = url
                                    cell.font = link_font
            
            # Auto-adjust column widths
            for column in ws.columns:
                max_length = 0
                column_letter = get_column_letter(column[0].column)
                
                for cell in column:
                    try:
                        # Get cell value for width calculation
                        cell_value = str(cell.value) if cell.value else ''
                        if len(cell_value) > max_length:
                            max_length = len(cell_value)
                    except:
                        pass
                
                # Set column width (with some padding, max 50)
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width
            
            # Set row height for header
            ws.row_dimensions[1].height = 20
            
            # Freeze header row
            ws.freeze_panes = 'A2'
            
            wb.save(filename)
        except Exception as e:
            # If formatting fails, the file still exists with data
            pass

